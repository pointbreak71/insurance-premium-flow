"""
Suyana — Colombia Parametric MVP Pilot (Synthetic Rainfall)
Full pipeline: land use, synthetic rainfall, gamma fitting, payouts, MVP optimisation.
"""

import time
import warnings
import traceback
from datetime import date, datetime
import numpy as np
import pandas as pd
import geopandas as gpd
from shapely.geometry import box, Point
from scipy import stats
from scipy.stats import gamma as gamma_dist, ks_2samp
from sklearn.covariance import LedoitWolf
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.colors as mcolors
import cvxpy as cp
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import ScatterChart, Reference, Series
import requests
import os
import pathlib

warnings.filterwarnings('ignore')

# ── paths ──────────────────────────────────────────────────────────────────────
ROOT = pathlib.Path(__file__).parent.parent
MAP_DIR   = ROOT / "outputs" / "maps"
EXCEL_DIR = ROOT / "outputs" / "excel"
DATA_DIR  = ROOT / "data" / "processed"
for d in [MAP_DIR, EXCEL_DIR, DATA_DIR]:
    d.mkdir(parents=True, exist_ok=True)

TODAY     = date.today()
XLSX_PATH = EXCEL_DIR / f"suyana_colombia_synthetic_mvp_{TODAY:%Y%m%d}.xlsx"

FLAGS      = []
STEP_LOG   = []
t0_global  = time.time()

def log(msg):
    elapsed = time.time() - t0_global
    print(f"[{elapsed:6.1f}s] {msg}")
    STEP_LOG.append(msg)

def flag(msg):
    FLAGS.append(msg)
    log(f"FLAG: {msg}")

np.random.seed(42)

# ══════════════════════════════════════════════════════════════════════════════
# PART A — LAND USE, CROP TYPES, GROWING SEASONS
# ══════════════════════════════════════════════════════════════════════════════

# ── A1 — Colombia boundary and grid ───────────────────────────────────────────
log("A1 — downloading Colombia boundary …")
try:
    colombia = gpd.read_file(
        "https://raw.githubusercontent.com/datasets/geo-countries/master/data/countries.geojson"
    )
    colombia = colombia[colombia['ADMIN'] == 'Colombia'].to_crs("EPSG:4326")
    log(f"A1 — boundary loaded, bounds: {colombia.total_bounds}")
except Exception as e:
    flag(f"A1 boundary download failed ({e}); using bounding box fallback")
    from shapely.geometry import Polygon
    # Colombia rough polygon fallback
    colombia = gpd.GeoDataFrame(
        {'ADMIN': ['Colombia']},
        geometry=[Polygon([(-79,1),(-67,1),(-67,12),(-79,12),(-79,1)])],
        crs="EPSG:4326"
    )

# Build 0.1° grid
RES = 0.1
lon_min, lat_min, lon_max, lat_max = -79.0, 0.0, -67.0, 12.0
lons = np.arange(lon_min, lon_max, RES)
lats = np.arange(lat_min, lat_max, RES)

cells = []
for lat in lats:
    for lon in lons:
        centroid = Point(lon + RES/2, lat + RES/2)
        cell_box = box(lon, lat, lon + RES, lat + RES)
        cells.append({'centroid_lon': lon + RES/2,
                      'centroid_lat': lat + RES/2,
                      'geometry': cell_box})

grid = gpd.GeoDataFrame(cells, crs="EPSG:4326")
# Clip to Colombia
colombia_union = colombia.geometry.union_all() if hasattr(colombia.geometry, 'union_all') else colombia.geometry.unary_union
mask = grid['geometry'].centroid.within(colombia_union)
grid = grid[mask].reset_index(drop=True)
grid['grid_id'] = ['G%04d' % i for i in range(len(grid))]

# Approximate area: at equator 0.1° ≈ 11.1 km, area ~ 11.1×11.1 km² = ~123.21 km² = 12321 ha
# At latitude φ: lon_km = 111.32 × cos(φ), lat_km = 110.57
# Use centroid lat
grid['cell_area_ha'] = (111320 * np.cos(np.radians(grid['centroid_lat'])) * RES / 1000) * \
                        (110570 * RES / 1000) * 100   # km² → ha

log(f"A1 — grid has {len(grid)} cells within Colombia")

# MAP 1
log("MAP1 — drawing Colombia grid …")
fig, ax = plt.subplots(figsize=(10, 12))
colombia.boundary.plot(ax=ax, color='black', linewidth=1.5)
grid.plot(ax=ax, facecolor='white', edgecolor='grey', linewidth=0.3, alpha=0.8)
ax.set_title("Map 1 — Colombia 0.1° Grid", fontsize=14, fontweight='bold')
ax.set_xlabel("Longitude"); ax.set_ylabel("Latitude")
ax.set_xlim(-80, -66); ax.set_ylim(-1, 13)
plt.tight_layout()
plt.savefig(MAP_DIR / "map1_colombia_grid.png", dpi=150)
plt.close()
log("MAP1 saved")

# ── A2 — Cropland extent and crop types ───────────────────────────────────────
log("A2 — assigning cropland zones (synthetic fallback) …")

ZONES = {
    'A': dict(lat_min=3, lat_max=7,  lon_min=-73, lon_max=-68,
              crops={'maize':0.60, 'rice':0.40}, frac=0.25, penetration=0.03),
    'B': dict(lat_min=3, lat_max=5,  lon_min=-77, lon_max=-75,
              crops={'sugarcane':0.55, 'maize':0.45}, frac=0.45, penetration=0.08),
    'C': dict(lat_min=8, lat_max=11, lon_min=-76, lon_max=-73,
              crops={'maize':0.50, 'sorghum':0.30, 'cotton':0.20}, frac=0.30, penetration=0.04),
    'D': dict(lat_min=1, lat_max=4,  lon_min=-76, lon_max=-73,
              crops={'coffee':0.50, 'maize':0.30, 'potato':0.20}, frac=0.20, penetration=0.03),
    'E': dict(lat_min=5, lat_max=9,  lon_min=-75, lon_max=-73,
              crops={'rice':0.45, 'maize':0.35, 'sorghum':0.20}, frac=0.25, penetration=0.04),
}

def assign_zone(lat, lon):
    for zname, z in ZONES.items():
        if z['lat_min'] <= lat < z['lat_max'] and z['lon_min'] <= lon < z['lon_max']:
            return zname
    return None

def primary_crop(zone):
    if zone is None:
        return None
    return max(ZONES[zone]['crops'], key=ZONES[zone]['crops'].get)

grid['zone'] = [assign_zone(r.centroid_lat, r.centroid_lon) for _, r in grid.iterrows()]
def is_valid_zone(z):
    return z is not None and not (isinstance(z, float) and np.isnan(z))

grid['is_cropland'] = grid['zone'].apply(is_valid_zone)
grid['primary_crop'] = grid['zone'].apply(lambda z: primary_crop(z) if is_valid_zone(z) else None)
grid['cropland_ha'] = grid.apply(
    lambda r: r.cell_area_ha * ZONES[r.zone]['frac'] if is_valid_zone(r.zone) else 0.0, axis=1)
grid['penetration_rate'] = grid['zone'].apply(
    lambda z: ZONES[z]['penetration'] if is_valid_zone(z) else 0.0)
grid['insured_ha']   = grid['cropland_ha'] * grid['penetration_rate']
grid['uninsured_ha'] = grid['cropland_ha'] * (1 - grid['penetration_rate'])

n_crop = grid['is_cropland'].sum()
log(f"A2 — {n_crop} cropland cells, total cropland ha: {grid['cropland_ha'].sum():,.0f}")

CROP_COLORS = {
    'maize':    '#FFA500',
    'rice':     '#4472C4',
    'sugarcane':'#70AD47',
    'sorghum':  '#8B4513',
    'cotton':   '#FF69B4',
    'coffee':   '#8B0000',
    'potato':   '#7030A0',
}

# MAP 2 — Cropland Extent
log("MAP2 — cropland extent …")
fig, ax = plt.subplots(figsize=(10, 12))
colombia.boundary.plot(ax=ax, color='black', linewidth=1.5, zorder=3)
non_crop = grid[~grid['is_cropland']]
non_crop.plot(ax=ax, color='#EEEEEE', edgecolor='none', zorder=1)
crop_grid = grid[grid['is_cropland']].copy()
norm = mcolors.Normalize(vmin=0, vmax=crop_grid["cropland_ha"].max())
cmap = plt.cm.Greens
crop_plot = crop_grid.copy()
crop_plot.plot(ax=ax, column='cropland_ha', cmap='Greens', norm=norm,
               edgecolor='none', zorder=2)
sm = plt.cm.ScalarMappable(cmap='Greens', norm=norm)
sm.set_array([])
plt.colorbar(sm, ax=ax, label='Cropland ha', shrink=0.6)
ax.set_title("Map 2 — Cropland Extent", fontsize=14, fontweight='bold')
ax.set_xlabel("Longitude"); ax.set_ylabel("Latitude")
ax.set_xlim(-80, -66); ax.set_ylim(-1, 13)
plt.tight_layout()
plt.savefig(MAP_DIR / "map2_cropland.png", dpi=150)
plt.close()
log("MAP2 saved")

# MAP 3 — Crop Types
log("MAP3 — crop types …")
fig, ax = plt.subplots(figsize=(10, 12))
colombia.boundary.plot(ax=ax, color='black', linewidth=1.5, zorder=3)
non_crop.plot(ax=ax, color='#EEEEEE', edgecolor='none', zorder=1)
for crop_name, color in CROP_COLORS.items():
    subset = grid[grid['primary_crop'] == crop_name]
    if len(subset):
        subset.plot(ax=ax, color=color, edgecolor='none', zorder=2, label=crop_name)
ax.legend(loc='lower right', fontsize=9)
ax.set_title("Map 3 — Crop Types (Primary)", fontsize=14, fontweight='bold')
ax.set_xlabel("Longitude"); ax.set_ylabel("Latitude")
ax.set_xlim(-80, -66); ax.set_ylim(-1, 13)
plt.tight_layout()
plt.savefig(MAP_DIR / "map3_crop_types.png", dpi=150)
plt.close()
log("MAP3 saved")

# ── A3 — Growing seasons and plots ────────────────────────────────────────────
log("A3 — assigning growing seasons …")

MAX_PAYOUT = {
    'maize':280, 'rice':320, 'sugarcane':350, 'sorghum':240,
    'cotton':300, 'coffee':420, 'potato':380
}

SEASONS = [
    dict(season_id='S1', season_name='Primera',        season_months='Mar-Jun',
         zones=['B','C','E'], crops=['maize','sorghum','cotton','rice']),
    dict(season_id='S2', season_name='Segunda',        season_months='Aug-Nov',
         zones=['B','C','E'], crops=['maize','sorghum','sugarcane']),
    dict(season_id='S3', season_name='Llanos wet',     season_months='Apr-Oct',
         zones=['A'],       crops=['maize','rice']),
    dict(season_id='S4', season_name='Andean main',    season_months='Mar-Jun',
         zones=['D'],       crops=['coffee','maize','potato']),
]

plots = []
for _, cell in grid[grid['is_cropland']].iterrows():
    zone = cell.zone
    crop = cell.primary_crop
    for s in SEASONS:
        if zone in s['zones'] and crop in s['crops']:
            plots.append({
                'grid_id':            cell.grid_id,
                'centroid_lat':       cell.centroid_lat,
                'centroid_lon':       cell.centroid_lon,
                'zone':               zone,
                'season_id':          s['season_id'],
                'season_name':        s['season_name'],
                'season_months':      s['season_months'],
                'crop_type':          crop,
                'uninsured_ha':       cell.uninsured_ha,
                'max_payout_usd_per_ha': MAX_PAYOUT[crop],
            })

plots_df = pd.DataFrame(plots)
plots_df.insert(0, 'plot_id', ['P%05d' % i for i in range(len(plots_df))])
n_plots = len(plots_df)
log(f"A3 — {n_plots} plots created")

# MAP 4 — Growing Seasons
log("MAP4 — growing seasons …")
season_colors = ['#E91E63','#2196F3','#FF9800','#4CAF50']
fig, axes = plt.subplots(2, 2, figsize=(14, 14))
axes = axes.flatten()
for idx, s in enumerate(SEASONS):
    ax = axes[idx]
    colombia.boundary.plot(ax=ax, color='black', linewidth=1.2, zorder=3)
    non_crop.plot(ax=ax, color='#EEEEEE', edgecolor='none', zorder=1)
    # Background cropland
    crop_grid.plot(ax=ax, color='#CCCCCC', edgecolor='none', zorder=1)
    # Season-specific cells
    season_plots = plots_df[plots_df['season_id'] == s['season_id']]
    season_gids  = season_plots['grid_id'].unique()
    season_cells = grid[grid['grid_id'].isin(season_gids)]
    season_cells.plot(ax=ax, color=season_colors[idx], edgecolor='none', zorder=2, alpha=0.8)
    ax.set_title(f"{s['season_id']} — {s['season_name']}\n({s['season_months']})",
                 fontsize=11, fontweight='bold')
    ax.set_xlim(-80, -66); ax.set_ylim(-1, 13)
    ax.set_xlabel("Lon"); ax.set_ylabel("Lat")
    n = len(season_plots)
    ax.text(0.02, 0.02, f"{n} plots", transform=ax.transAxes, fontsize=9, color='#333333')
plt.suptitle("Map 4 — Growing Seasons", fontsize=14, fontweight='bold', y=1.01)
plt.tight_layout()
plt.savefig(MAP_DIR / "map4_growing_seasons.png", dpi=150, bbox_inches='tight')
plt.close()
log("MAP4 saved")

# ══════════════════════════════════════════════════════════════════════════════
# PART B — SYNTHETIC RAINFALL
# ══════════════════════════════════════════════════════════════════════════════

# ── B1 — Simulate 30 years ────────────────────────────────────────────────────
log("B1 — simulating 30 years of synthetic rainfall …")

YEARS = list(range(1994, 2024))
N_YEARS = len(YEARS)

GAMMA_PARAMS = {
    ('A','S3'): dict(mean=1850, cv=0.22),
    ('B','S1'): dict(mean=920,  cv=0.28),
    ('B','S2'): dict(mean=880,  cv=0.30),
    ('C','S1'): dict(mean=680,  cv=0.32),
    ('C','S2'): dict(mean=620,  cv=0.35),
    ('D','S4'): dict(mean=750,  cv=0.25),
    ('E','S1'): dict(mean=800,  cv=0.29),
    ('E','S2'): dict(mean=760,  cv=0.31),
}

ENSO_YEARS = {
    'el_nino': {1997,1998,2002,2003,2009,2010,2015,2016,2019},
    'la_nina': {1995,1996,1999,2000,2007,2008,2010,2011,2020,2021,2022},
}
def enso_phase(yr):
    if yr in ENSO_YEARS['el_nino']:  return 'El Niño'
    if yr in ENSO_YEARS['la_nina']:  return 'La Niña'
    return 'Neutral'

ENSO_MOD = {
    'El Niño': {'A':0.75,'B':1.10,'C':0.75,'D':1.10,'E':0.75},
    'La Niña': {'A':1.20,'B':0.85,'C':1.20,'D':0.85,'E':1.20},
    'Neutral':  {z:1.00 for z in 'ABCDE'},
}

# Common zone-year shocks
zone_year_shocks = {}
for z in 'ABCDE':
    for yr in YEARS:
        # Pick the first matching gamma params for this zone
        matching = [v for (zz,ss), v in GAMMA_PARAMS.items() if zz == z]
        if matching:
            zone_mean = matching[0]['mean']
        else:
            zone_mean = 800
        zone_year_shocks[(z, yr)] = np.random.normal(0, 0.08 * zone_mean)

rainfall = np.zeros((n_plots, N_YEARS))

for pidx, plot in plots_df.iterrows():
    z  = plot.zone
    sid = plot.season_id
    key = (z, sid)
    if key not in GAMMA_PARAMS:
        # fallback
        params = dict(mean=700, cv=0.30)
        flag(f"B1 missing gamma params for zone={z} season={sid}; using fallback")
    else:
        params = GAMMA_PARAMS[key]

    alpha = 1.0 / params['cv']**2
    for yidx, yr in enumerate(YEARS):
        phase = enso_phase(yr)
        mod   = ENSO_MOD[phase][z]
        mean_adj = params['mean'] * mod
        beta_adj = mean_adj / alpha
        shock = zone_year_shocks[(z, yr)]
        raw = np.random.gamma(alpha, beta_adj) + shock
        rainfall[pidx, yidx] = max(10.0, raw)

rain_df = pd.DataFrame(rainfall, index=plots_df['plot_id'], columns=YEARS)
rain_df['mean_mm'] = rain_df[YEARS].mean(axis=1)
rain_df['std_mm']  = rain_df[YEARS].std(axis=1)
rain_df['cv']      = rain_df['std_mm'] / rain_df['mean_mm']
rain_df['min_mm']  = rain_df[YEARS].min(axis=1)
rain_df['max_mm']  = rain_df[YEARS].max(axis=1)

log(f"B1 — rainfall matrix shape: {n_plots}×{N_YEARS}; mean across all: {rain_df['mean_mm'].mean():.0f} mm")

# ── B2 — Gamma fit and percentiles ───────────────────────────────────────────
log("B2 — fitting gamma distributions …")

fit_rows = []
percentiles_matrix = np.zeros((n_plots, 4))  # p1, p10, p90, p99

for pidx, plot in plots_df.iterrows():
    series = rainfall[pidx, :]
    try:
        a_fit, loc_fit, scale_fit = gamma_dist.fit(series, floc=0)
        # KS test
        ks_stat, ks_p = stats.kstest(series, 'gamma', args=(a_fit, loc_fit, scale_fit))
        if ks_p < 0.05:
            method = 'empirical'
            p1  = np.percentile(series, 1)
            p10 = np.percentile(series, 10)
            p90 = np.percentile(series, 90)
            p99 = np.percentile(series, 99)
        else:
            method = 'gamma'
            p1  = gamma_dist.ppf(0.01, a_fit, loc_fit, scale_fit)
            p10 = gamma_dist.ppf(0.10, a_fit, loc_fit, scale_fit)
            p90 = gamma_dist.ppf(0.90, a_fit, loc_fit, scale_fit)
            p99 = gamma_dist.ppf(0.99, a_fit, loc_fit, scale_fit)
        beta_fit = scale_fit
    except Exception as e:
        flag(f"B2 fit failed for plot {plot.plot_id}: {e}")
        a_fit = loc_fit = 0; scale_fit = 1; ks_p = 0; method = 'empirical'
        p1  = np.percentile(series, 1)
        p10 = np.percentile(series, 10)
        p90 = np.percentile(series, 90)
        p99 = np.percentile(series, 99)
        beta_fit = scale_fit

    percentiles_matrix[pidx] = [p1, p10, p90, p99]
    fit_rows.append({
        'plot_id':    plot.plot_id,
        'zone':       plot.zone,
        'season_id':  plot.season_id,
        'alpha':      round(a_fit, 4),
        'beta':       round(beta_fit, 4),
        'p1_mm':      round(p1, 1),
        'p10_mm':     round(p10, 1),
        'p90_mm':     round(p90, 1),
        'p99_mm':     round(p99, 1),
        'ks_pvalue':  round(ks_p, 4),
        'fit_method': method,
    })

fit_df = pd.DataFrame(fit_rows)
log(f"B2 — fit complete; empirical fallback count: {(fit_df.fit_method=='empirical').sum()}")

# ── B3 — Payout schedule ──────────────────────────────────────────────────────
log("B3 — computing payout percentages …")

def compute_payout(rain_val, p10, p1, p90, p99):
    if rain_val <= p1:
        deficit = 1.0
    elif rain_val < p10:
        deficit = (p10 - rain_val) / (p10 - p1) if (p10 - p1) > 0 else 0.0
    else:
        deficit = 0.0

    if rain_val >= p99:
        excess = 1.0
    elif rain_val > p90:
        excess = (rain_val - p90) / (p99 - p90) if (p99 - p90) > 0 else 0.0
    else:
        excess = 0.0

    return deficit, excess

deficit_pct = np.zeros((n_plots, N_YEARS))
excess_pct  = np.zeros((n_plots, N_YEARS))

for pidx in range(n_plots):
    p1, p10, p90, p99 = percentiles_matrix[pidx]
    for yidx in range(N_YEARS):
        d, e = compute_payout(rainfall[pidx, yidx], p10, p1, p90, p99)
        deficit_pct[pidx, yidx] = d
        excess_pct[pidx, yidx]  = e

deficit_df = pd.DataFrame(deficit_pct, index=plots_df['plot_id'], columns=YEARS)
excess_df  = pd.DataFrame(excess_pct,  index=plots_df['plot_id'], columns=YEARS)

deficit_df['trigger_freq_pct'] = (deficit_pct > 0).mean(axis=1) * 100
excess_df['trigger_freq_pct']  = (excess_pct  > 0).mean(axis=1) * 100

log(f"B3 — deficit trigger freq: {deficit_df['trigger_freq_pct'].mean():.1f}%, "
    f"excess: {excess_df['trigger_freq_pct'].mean():.1f}%")

# ── B4 — USD payout history ───────────────────────────────────────────────────
log("B4 — computing USD payouts …")

max_payout_arr = plots_df['max_payout_usd_per_ha'].values
uninsured_arr  = plots_df['uninsured_ha'].values
max_usd_per_plot = max_payout_arr * uninsured_arr   # shape (n_plots,)

usd_deficit = deficit_pct * max_usd_per_plot[:, None]
usd_excess  = excess_pct  * max_usd_per_plot[:, None]

usd_def_df = pd.DataFrame(usd_deficit, index=plots_df['plot_id'], columns=YEARS)
usd_exc_df = pd.DataFrame(usd_excess,  index=plots_df['plot_id'], columns=YEARS)

for df in [usd_def_df, usd_exc_df]:
    df['mean_annual_payout']  = df[YEARS].mean(axis=1)
    df['std_annual_payout']   = df[YEARS].std(axis=1)
    df['max_year_payout']     = df[YEARS].max(axis=1)
    df['total_30yr_payout']   = df[YEARS].sum(axis=1)
    df['trigger_frequency_pct'] = (df[YEARS] > 0).mean(axis=1) * 100

total_exposure = max_usd_per_plot.sum()
log(f"B4 — total max exposure: ${total_exposure:,.0f}  |  "
    f"mean annual deficit payout: ${usd_def_df['mean_annual_payout'].sum():,.0f}")

# ══════════════════════════════════════════════════════════════════════════════
# PART C — PORTFOLIO OPTIMISATION
# ══════════════════════════════════════════════════════════════════════════════

# ── C1 — Aggregate to zone-season-peril slots ─────────────────────────────────
log("C1 — aggregating to portfolio slots …")

plots_meta = plots_df[['plot_id','zone','season_id']].set_index('plot_id')

def build_slots(usd_matrix_df, peril_label):
    usd = usd_matrix_df[YEARS].copy()
    usd['zone']      = plots_meta['zone']
    usd['season_id'] = plots_meta['season_id']
    slot_annual = usd.groupby(['zone','season_id'])[YEARS].sum()
    slot_annual.index = [f"{z}_{s}_{peril_label}" for z, s in slot_annual.index]
    return slot_annual

slot_deficit = build_slots(usd_def_df, 'DEF')
slot_excess  = build_slots(usd_exc_df, 'EXC')
slots_all    = pd.concat([slot_deficit, slot_excess])  # (n_slots, N_YEARS)

# Remove zero slots
slots_all = slots_all[slots_all[YEARS].sum(axis=1) > 0]
n_slots = len(slots_all)
log(f"C1 — {n_slots} non-zero portfolio slots")

total_portfolio = slots_all[YEARS].sum(axis=0)
slot_summary = slots_all.copy()
slot_summary['Mean']   = slots_all[YEARS].mean(axis=1)
slot_summary['Std']    = slots_all[YEARS].std(axis=1)
slot_summary['P90']    = slots_all[YEARS].quantile(0.90, axis=1)
slot_summary['Max']    = slots_all[YEARS].max(axis=1)

# ── C2 — Correlation and covariance ──────────────────────────────────────────
log("C2 — computing Spearman correlation and Ledoit-Wolf covariance …")

X = slots_all[YEARS].values.T  # (N_YEARS, n_slots)
spearman_corr = pd.DataFrame(
    stats.spearmanr(X).statistic if n_slots > 1 else np.array([[1.0]]),
    index=slots_all.index, columns=slots_all.index
)

lw = LedoitWolf()
lw.fit(X)
Sigma = lw.covariance_
shrinkage = lw.shrinkage_

# Regularise
min_eig = np.linalg.eigvalsh(Sigma).min()
if min_eig < 1e-8:
    Sigma += (abs(min_eig) + 1e-6) * np.eye(n_slots)
    flag("C2 covariance regularised (negative eigenvalue)")

cond_number = np.linalg.cond(Sigma)
mu = slots_all[YEARS].mean(axis=1).values  # expected annual payout per slot
log(f"C2 — shrinkage={shrinkage:.3f}, condition number={cond_number:.1f}")

# ── C3 — Minimum variance portfolio ──────────────────────────────────────────
log("C3 — solving MVP optimisation …")

def solve_mvp(Sigma, mu, constrained=False):
    n = len(mu)
    # Use symmetric positive-definite matrix
    Sigma_psd = (Sigma + Sigma.T) / 2 + 1e-6 * np.eye(n)
    w = cp.Variable(n)
    objective = cp.Minimize(cp.quad_form(w, cp.Parameter(shape=(n,n), value=Sigma_psd, PSD=True)))
    constraints = [cp.sum(w) == 1, w >= 0]
    if constrained:
        constraints += [w <= 0.25]
    prob = cp.Problem(objective, constraints)
    for solver in [cp.SCS, cp.ECOS, None]:
        try:
            if solver:
                prob.solve(solver=solver, verbose=False)
            else:
                prob.solve(verbose=False)
            if w.value is not None:
                break
        except Exception:
            continue
    if w.value is None:
        flag("C3 solver returned None; using uniform weights")
        return np.ones(n) / n
    wv = np.clip(w.value, 0, None)
    total = wv.sum()
    return wv / total if total > 0 else np.ones(n) / n

w_unc = solve_mvp(Sigma, mu, constrained=False)
w_con = solve_mvp(Sigma, mu, constrained=True)

def portfolio_stats(w, Sigma, mu):
    var = w @ Sigma @ w
    std = np.sqrt(var)
    exp = w @ mu
    premium = exp * 1.3
    # Diversification ratio: weighted avg vol / portfolio vol
    vols = np.sqrt(np.diag(Sigma))
    dr   = (w @ vols) / std if std > 0 else 1.0
    return dict(variance=var, std_dev=std, expected_payout=exp,
                implied_premium=premium, diversification_ratio=dr)

stats_unc = portfolio_stats(w_unc, Sigma, mu)
stats_con = portfolio_stats(w_con, Sigma, mu)

log(f"C3 — Unconstrained MVP: std=${stats_unc['std_dev']:,.0f}, DR={stats_unc['diversification_ratio']:.2f}")
log(f"C3 — Constrained   MVP: std=${stats_con['std_dev']:,.0f}, DR={stats_con['diversification_ratio']:.2f}")

# Efficient frontier — 20 points
mu_min, mu_max = mu.min(), mu.max()
frontier_targets = np.linspace(mu_min, mu_max, 20)
frontier_rows = []
Sigma_psd_f = (Sigma + Sigma.T) / 2 + 1e-6 * np.eye(len(mu))
for target in frontier_targets:
    n = len(mu)
    w = cp.Variable(n)
    Sp = cp.Parameter(shape=(n,n), value=Sigma_psd_f, PSD=True)
    objective = cp.Minimize(cp.quad_form(w, Sp))
    constraints = [cp.sum(w) == 1, w >= 0, w @ mu >= target]
    prob = cp.Problem(objective, constraints)
    solved = False
    for solver in [cp.SCS, cp.ECOS, None]:
        try:
            if solver:
                prob.solve(solver=solver, verbose=False)
            else:
                prob.solve(verbose=False)
            if w.value is not None:
                solved = True
                break
        except Exception:
            continue
    if w.value is not None:
        wv = np.clip(w.value, 0, None)
        wv /= wv.sum()
        var = wv @ Sigma @ wv
        frontier_rows.append({
            'target_payout': target,
            'portfolio_std':  np.sqrt(var),
            'expected_payout': wv @ mu,
            'implied_premium': (wv @ mu) * 1.3,
        })

frontier_df = pd.DataFrame(frontier_rows)

# Auto-generate MVP summary
top_unc = pd.Series(w_unc, index=slots_all.index).nlargest(3)
top_con = pd.Series(w_con, index=slots_all.index).nlargest(3)

# Correlation insights
flat_corr = spearman_corr.values.copy()
np.fill_diagonal(flat_corr, np.nan)
if n_slots > 1:
    min_corr_idx = np.unravel_index(np.nanargmin(flat_corr), flat_corr.shape)
    neg_pair = (spearman_corr.index[min_corr_idx[0]], spearman_corr.columns[min_corr_idx[1]])
    neg_corr_val = flat_corr[min_corr_idx]
else:
    neg_pair = ("N/A", "N/A"); neg_corr_val = 0

mvp_summary = (
    f"UNCONSTRAINED MVP: Top slots — {', '.join(f'{s} ({w:.1%})' for s,w in top_unc.items())}. "
    f"Portfolio std dev: ${stats_unc['std_dev']:,.0f}/yr. "
    f"Expected annual payout: ${stats_unc['expected_payout']:,.0f}. "
    f"Implied premium at 1.3× LR: ${stats_unc['implied_premium']:,.0f}. "
    f"Diversification ratio: {stats_unc['diversification_ratio']:.2f}. "
    f"\nCONSTRAINED MVP (≤25%): Top slots — {', '.join(f'{s} ({w:.1%})' for s,w in top_con.items())}. "
    f"Portfolio std dev: ${stats_con['std_dev']:,.0f}/yr. "
    f"Most negatively correlated pair: {neg_pair[0]} vs {neg_pair[1]} (ρ={neg_corr_val:.2f}). "
    f"ENSO implication: El Niño reduces deficit risk in Zones B/D but amplifies it in A/C/E; "
    f"La Niña has the reverse effect — blending bimodal zones (B,C,E) with Llanos (A) provides "
    f"partial ENSO hedging."
)
log(f"C3 — summary generated ({len(mvp_summary)} chars)")

# ── C4 — ENSO stress test ─────────────────────────────────────────────────────
log("C4 — ENSO stress test …")

slot_annual_vals = slots_all[YEARS].values  # (n_slots, N_YEARS)
w_stress = w_con

stress_rows = []
for yidx, yr in enumerate(YEARS):
    payout_per_slot = slot_annual_vals[:, yidx] * w_stress
    total = payout_per_slot.sum()
    mean_all = (slot_annual_vals * w_stress[:, None]).sum(axis=0).mean()
    max_all  = (slot_annual_vals * w_stress[:, None]).sum(axis=0).max()
    row = {'year': yr, 'enso_phase': enso_phase(yr), 'total_payout': total,
           'pct_of_mean': total/mean_all*100 if mean_all > 0 else 0,
           'pct_of_max':  total/max_all*100  if max_all  > 0 else 0}
    for i, slot_name in enumerate(slots_all.index):
        row[slot_name] = payout_per_slot[i]
    stress_rows.append(row)

stress_df = pd.DataFrame(stress_rows).set_index('year')
enso_means = stress_df.groupby('enso_phase')['total_payout'].mean()
log(f"C4 — ENSO payout means: {enso_means.to_dict()}")

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
log("EXCEL — building workbook …")

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── helpers ───────────────────────────────────────────────────────────────────
BLUE_TAB   = "4472C4"
GREEN_TAB  = "70AD47"
WHITE_TAB  = "FFFFFF"

HDR_FILL   = PatternFill("solid", fgColor="2E4057")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT  = Font(bold=True)
USD_FMT    = '$#,##0'
PCT_FMT    = '0.0%'
NUM2_FMT   = '0.00'
CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left')

def style_header_row(ws, row=1, cols=None):
    max_col = ws.max_column if cols is None else cols
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER

def auto_width(ws, min_w=8, max_w=30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))

def freeze_and_bold(ws):
    ws.freeze_panes = 'A2'
    style_header_row(ws)

def write_df_to_sheet(ws, df, start_row=1, index=True, header=True):
    if header:
        cols = ([df.index.name or 'index'] if index else []) + list(df.columns)
        for ci, c in enumerate(cols, 1):
            ws.cell(row=start_row, column=ci, value=str(c))
    data_start = start_row + (1 if header else 0)
    for ri, (idx, row) in enumerate(df.iterrows()):
        row_num = data_start + ri
        col_offset = 1
        if index:
            ws.cell(row=row_num, column=1, value=idx)
            col_offset = 2
        for ci, val in enumerate(row, col_offset):
            ws.cell(row=row_num, column=ci, value=val if not (isinstance(val, float) and np.isnan(val)) else None)

# ── Cover sheet ───────────────────────────────────────────────────────────────
ws_cover = wb.create_sheet("Cover")
ws_cover.sheet_properties.tabColor = WHITE_TAB

title_font = Font(bold=True, size=16, color="2E4057")
ws_cover['B2'] = "Suyana — Colombia Parametric MVP Pilot (Synthetic Rainfall)"
ws_cover['B2'].font = title_font
ws_cover['B3'] = f"Generated: {datetime.now():%Y-%m-%d %H:%M}"
ws_cover['B3'].font = Font(italic=True, color="666666")

info_rows = [
    ("",""),
    ("DATA SOURCES",""),
    ("Land use:",    "Synthetic (SPAM2017 download not attempted; synthetic zones used)"),
    ("Rainfall:",    "SYNTHETIC — gamma-distributed, ENSO-correlated. NOT real CHIRPS data."),
    ("Boundary:",    "Natural Earth countries GeoJSON"),
    ("",""),
    ("GRID & COVERAGE",""),
    ("Grid resolution:", "0.1° (~11 km)"),
    ("Grid cells (Colombia):", f"{len(grid):,}"),
    ("Cropland cells:", f"{n_crop:,}"),
    ("Total plots:", f"{n_plots:,}"),
    ("Portfolio slots:", f"{n_slots:,}"),
    ("Years simulated:", f"{YEARS[0]}–{YEARS[-1]} ({N_YEARS} years)"),
    ("Total uninsured ha:", f"{grid['uninsured_ha'].sum():,.0f}"),
    ("Total max USD exposure:", f"${total_exposure:,.0f}"),
    ("",""),
    ("MVP SUMMARY",""),
]
for r, (k, v) in enumerate(info_rows, 5):
    ws_cover.cell(row=r, column=2, value=k).font = BOLD_FONT if k.isupper() else Font()
    ws_cover.cell(row=r, column=3, value=v)

# MVP summary paragraph
mvp_row = 5 + len(info_rows)
ws_cover.cell(row=mvp_row, column=2, value="Portfolio Analysis:").font = BOLD_FONT
ws_cover.merge_cells(start_row=mvp_row+1, start_column=2, end_row=mvp_row+6, end_column=6)
cell = ws_cover.cell(row=mvp_row+1, column=2, value=mvp_summary)
cell.alignment = Alignment(wrap_text=True, vertical='top')

# ToC
toc_row = mvp_row + 9
ws_cover.cell(row=toc_row, column=2, value="TABLE OF CONTENTS").font = Font(bold=True, size=12)
toc = [
    ("Cover",          "This sheet — summary and metadata"),
    ("A2_Exposure",    "Grid cells, zones, crop types, uninsured area"),
    ("A3_Plots",       "Plot definitions: grid × season × crop"),
    ("B1_SyntheticRainfall", "30-year synthetic seasonal rainfall (mm)"),
    ("B2_GammaFit",    "Gamma distribution parameters and percentile thresholds"),
    ("B3_PayoutPct",   "Deficit and excess payout percentages by plot-year"),
    ("B4_USDPayouts",  "USD payout history per plot"),
    ("C1_SlotAnnual",  "Aggregated portfolio slot payouts by year"),
    ("C2_Correlations","Spearman correlation matrix with Ledoit-Wolf shrinkage"),
    ("C3_MVPWeights",  "Minimum variance portfolio weights and efficient frontier"),
    ("C4_StressTest",  "ENSO stress test — year-by-year constrained MVP payouts"),
]
for i, (sname, desc) in enumerate(toc):
    ws_cover.cell(row=toc_row+1+i, column=2, value=sname).font = Font(bold=True, color="2E4057")
    ws_cover.cell(row=toc_row+1+i, column=3, value=desc)

ws_cover.column_dimensions['B'].width = 28
ws_cover.column_dimensions['C'].width = 60

# ── A2_Exposure ───────────────────────────────────────────────────────────────
ws_a2 = wb.create_sheet("A2_Exposure")
ws_a2.sheet_properties.tabColor = BLUE_TAB
exp_out = grid[['grid_id','centroid_lat','centroid_lon','zone','is_cropland',
                 'primary_crop','cell_area_ha','cropland_ha','insured_ha','uninsured_ha']].copy()
exp_out.columns = ['grid_id','lat','lon','zone','is_cropland','primary_crop',
                   'cell_area_ha','cropland_ha','insured_ha','uninsured_ha']
write_df_to_sheet(ws_a2, exp_out, index=False)
freeze_and_bold(ws_a2)
for row in ws_a2.iter_rows(min_row=2, min_col=8, max_col=10):
    for cell in row:
        cell.number_format = '#,##0'
auto_width(ws_a2)

# ── A3_Plots ──────────────────────────────────────────────────────────────────
ws_a3 = wb.create_sheet("A3_Plots")
ws_a3.sheet_properties.tabColor = BLUE_TAB
write_df_to_sheet(ws_a3, plots_df, index=False)
freeze_and_bold(ws_a3)
for row in ws_a3.iter_rows(min_row=2, min_col=plots_df.columns.get_loc('uninsured_ha')+1,
                            max_col=plots_df.columns.get_loc('uninsured_ha')+2):
    for cell in row:
        cell.number_format = '#,##0'
auto_width(ws_a3)

# ── B1_SyntheticRainfall ──────────────────────────────────────────────────────
ws_b1 = wb.create_sheet("B1_SyntheticRainfall")
ws_b1.sheet_properties.tabColor = BLUE_TAB
ws_b1['A1'] = "NOTE: SYNTHETIC DATA — not real CHIRPS."
ws_b1['A1'].font = Font(bold=True, color="FF0000")
write_df_to_sheet(ws_b1, rain_df, start_row=2)
style_header_row(ws_b1, row=2)
ws_b1.freeze_panes = 'B3'
# Color scale on rainfall values
max_row_b1 = 2 + n_plots
if n_plots > 0:
    cs_rule = ColorScaleRule(
        start_type='min', start_color='FF4444',
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max', end_color='4472C4'
    )
    # Apply to year columns only (cols 2 .. N_YEARS+1)
    end_col_b1 = get_column_letter(1 + N_YEARS)
    ws_b1.conditional_formatting.add(f'B3:{end_col_b1}{max_row_b1}', cs_rule)
auto_width(ws_b1, max_w=12)

# ── B2_GammaFit ───────────────────────────────────────────────────────────────
ws_b2 = wb.create_sheet("B2_GammaFit")
ws_b2.sheet_properties.tabColor = BLUE_TAB
write_df_to_sheet(ws_b2, fit_df, index=False)
freeze_and_bold(ws_b2)
auto_width(ws_b2)

# ── B3_PayoutPct ──────────────────────────────────────────────────────────────
ws_b3 = wb.create_sheet("B3_PayoutPct")
ws_b3.sheet_properties.tabColor = BLUE_TAB
ws_b3['A1'] = "DEFICIT PAYOUT %"
ws_b3['A1'].font = Font(bold=True, size=12)
write_df_to_sheet(ws_b3, deficit_df, start_row=2)
style_header_row(ws_b3, row=2)
ws_b3.freeze_panes = 'B3'

excess_start = n_plots + 5
ws_b3.cell(row=excess_start, column=1, value="EXCESS PAYOUT %").font = Font(bold=True, size=12)
write_df_to_sheet(ws_b3, excess_df, start_row=excess_start+1)
style_header_row(ws_b3, row=excess_start+1)

# format year cols as %
for row in ws_b3.iter_rows(min_row=3, max_row=2+n_plots, min_col=2, max_col=N_YEARS+1):
    for cell in row:
        cell.number_format = '0.0%'
auto_width(ws_b3, max_w=10)

# ── B4_USDPayouts ─────────────────────────────────────────────────────────────
ws_b4 = wb.create_sheet("B4_USDPayouts")
ws_b4.sheet_properties.tabColor = BLUE_TAB
ws_b4['A1'] = "DEFICIT USD PAYOUTS"
ws_b4['A1'].font = Font(bold=True, size=12)
write_df_to_sheet(ws_b4, usd_def_df, start_row=2)
style_header_row(ws_b4, row=2)
ws_b4.freeze_panes = 'B3'

excess_start_b4 = n_plots + 5
ws_b4.cell(row=excess_start_b4, column=1, value="EXCESS USD PAYOUTS").font = Font(bold=True, size=12)
write_df_to_sheet(ws_b4, usd_exc_df, start_row=excess_start_b4+1)
style_header_row(ws_b4, row=excess_start_b4+1)

for row in ws_b4.iter_rows(min_row=3, max_row=2+n_plots, min_col=2, max_col=N_YEARS+1):
    for cell in row:
        cell.number_format = USD_FMT
auto_width(ws_b4, max_w=14)

# ── C1_SlotAnnual ─────────────────────────────────────────────────────────────
ws_c1 = wb.create_sheet("C1_SlotAnnual")
ws_c1.sheet_properties.tabColor = GREEN_TAB

# Transpose: rows=years, cols=slots
slot_years_df = slots_all[YEARS].T.copy()
slot_years_df.index.name = 'year'
slot_years_df['Total'] = total_portfolio

write_df_to_sheet(ws_c1, slot_years_df)
freeze_and_bold(ws_c1)

# Summary rows
summary_start = N_YEARS + 3
labels = ['Mean','Std','P90','Max']
funcs  = [np.mean, np.std, lambda x: np.percentile(x, 90), np.max]
for li, (lbl, fn) in enumerate(zip(labels, funcs)):
    r = summary_start + li
    ws_c1.cell(row=r, column=1, value=lbl).font = BOLD_FONT
    for ci, col in enumerate(slot_years_df.columns, 2):
        vals = slot_years_df[col].values
        ws_c1.cell(row=r, column=ci, value=fn(vals)).number_format = USD_FMT

for row in ws_c1.iter_rows(min_row=2, max_row=N_YEARS+1, min_col=2):
    for cell in row:
        cell.number_format = USD_FMT
auto_width(ws_c1, max_w=16)

# ── C2_Correlations ────────────────────────────────────────────────────────────
ws_c2 = wb.create_sheet("C2_Correlations")
ws_c2.sheet_properties.tabColor = GREEN_TAB
ws_c2['A1'] = f"Spearman Correlation Matrix (Ledoit-Wolf shrinkage={shrinkage:.3f}, cond#{cond_number:.0f})"
ws_c2['A1'].font = Font(bold=True)

write_df_to_sheet(ws_c2, spearman_corr, start_row=2)
style_header_row(ws_c2, row=2)
ws_c2.freeze_panes = 'B3'

# Colour scale: dark blue -1 → white 0 → dark red 1
if n_slots > 1:
    end_col_c2 = get_column_letter(n_slots + 1)
    end_row_c2 = n_slots + 2
    cs_corr = ColorScaleRule(
        start_type='num', start_value=-1, start_color='4472C4',
        mid_type='num', mid_value=0,    mid_color='FFFFFF',
        end_type='num', end_value=1,    end_color='FF0000'
    )
    ws_c2.conditional_formatting.add(f'B3:{end_col_c2}{end_row_c2}', cs_corr)
for row in ws_c2.iter_rows(min_row=3, max_row=n_slots+2, min_col=2, max_col=n_slots+1):
    for cell in row:
        cell.number_format = '0.00'
auto_width(ws_c2, max_w=16)

# ── C3_MVPWeights ─────────────────────────────────────────────────────────────
ws_c3 = wb.create_sheet("C3_MVPWeights")
ws_c3.sheet_properties.tabColor = GREEN_TAB

ws_c3['A1'] = "Minimum Variance Portfolio Weights"
ws_c3['A1'].font = Font(bold=True, size=13)

headers = ['Slot','Unconstrained Weight','Constrained Weight (≤25%)',
           'Expected Annual Payout ($)','Max Payout ($)']
for ci, h in enumerate(headers, 1):
    c = ws_c3.cell(row=3, column=ci, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER

for ri, slot_name in enumerate(slots_all.index):
    r = 4 + ri
    ws_c3.cell(row=r, column=1, value=slot_name)
    ws_c3.cell(row=r, column=2, value=round(w_unc[ri], 6)).number_format = '0.0%'
    ws_c3.cell(row=r, column=3, value=round(w_con[ri], 6)).number_format = '0.0%'
    ws_c3.cell(row=r, column=4, value=round(mu[ri], 0)).number_format = USD_FMT
    ws_c3.cell(row=r, column=5, value=round(slots_all[YEARS].iloc[ri].max(), 0)).number_format = USD_FMT

# Highlight weights
yellow_fill = PatternFill("solid", fgColor="FFFF00")
orange_fill = PatternFill("solid", fgColor="FFA500")
for ri in range(n_slots):
    r = 4 + ri
    for ci in [2, 3]:
        cell = ws_c3.cell(row=r, column=ci)
        if cell.value and cell.value > 0.25:
            cell.fill = orange_fill
        elif cell.value and cell.value > 0.15:
            cell.fill = yellow_fill

# Stats rows
stats_start = 4 + n_slots + 2
labels_stats = ['Portfolio Variance','Portfolio Std Dev','Expected Annual Payout',
                'Implied Premium (1.3×)','Diversification Ratio']
unc_vals = [stats_unc['variance'], stats_unc['std_dev'], stats_unc['expected_payout'],
            stats_unc['implied_premium'], stats_unc['diversification_ratio']]
con_vals = [stats_con['variance'], stats_con['std_dev'], stats_con['expected_payout'],
            stats_con['implied_premium'], stats_con['diversification_ratio']]
for li, (lbl, uv, cv) in enumerate(zip(labels_stats, unc_vals, con_vals)):
    r = stats_start + li
    ws_c3.cell(row=r, column=1, value=lbl).font = BOLD_FONT
    ws_c3.cell(row=r, column=2, value=round(uv, 2))
    ws_c3.cell(row=r, column=3, value=round(cv, 2))

# Interpretation
interp_row = stats_start + len(labels_stats) + 2
ws_c3.cell(row=interp_row, column=1, value="INTERPRETATION").font = Font(bold=True, size=11)
ws_c3.merge_cells(start_row=interp_row+1, start_column=1, end_row=interp_row+8, end_column=5)
interp_cell = ws_c3.cell(row=interp_row+1, column=1, value=mvp_summary)
interp_cell.alignment = Alignment(wrap_text=True, vertical='top')

ws_c3.freeze_panes = 'A4'
auto_width(ws_c3, max_w=30)

# Efficient Frontier sheet continuation
ws_c3_ef = wb.create_sheet("C3_EfficientFrontier")
ws_c3_ef.sheet_properties.tabColor = GREEN_TAB
ws_c3_ef['A1'] = "Efficient Frontier — 20 Points"
ws_c3_ef['A1'].font = Font(bold=True, size=12)
write_df_to_sheet(ws_c3_ef, frontier_df.reset_index(drop=True), start_row=2, index=False)
freeze_and_bold(ws_c3_ef)
for row in ws_c3_ef.iter_rows(min_row=3, max_row=22, min_col=2, max_col=4):
    for cell in row:
        cell.number_format = USD_FMT
auto_width(ws_c3_ef)

# ── C4_StressTest ─────────────────────────────────────────────────────────────
ws_c4 = wb.create_sheet("C4_StressTest")
ws_c4.sheet_properties.tabColor = GREEN_TAB

stress_out = stress_df.reset_index()
write_df_to_sheet(ws_c4, stress_out, index=False)
freeze_and_bold(ws_c4)

# USD formatting on numeric cols
for row in ws_c4.iter_rows(min_row=2, max_row=N_YEARS+2, min_col=4):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = USD_FMT

# Color scale on Total Payout column (col 4, "total_payout")
total_col_idx = list(stress_out.columns).index('total_payout') + 1
total_col_ltr = get_column_letter(total_col_idx)
cs_stress = ColorScaleRule(
    start_type='min', start_color='FFFFFF',
    end_type='max', end_color='FF0000'
)
ws_c4.conditional_formatting.add(
    f'{total_col_ltr}2:{total_col_ltr}{N_YEARS+1}', cs_stress
)

# Bold red top 3 payout years
top3_years = stress_df['total_payout'].nlargest(3).index
for ri, yr in enumerate(YEARS):
    r = ri + 2
    if yr in top3_years:
        for ci in range(1, len(stress_out.columns)+1):
            cell = ws_c4.cell(row=r, column=ci)
            cell.font = Font(bold=True, color="FF0000")

# ENSO phase summary
enso_sum_row = N_YEARS + 4
ws_c4.cell(row=enso_sum_row, column=1, value="Mean Payout by ENSO Phase").font = BOLD_FONT
for pi, (phase, val) in enumerate(enso_means.items()):
    ws_c4.cell(row=enso_sum_row+1+pi, column=1, value=phase)
    ws_c4.cell(row=enso_sum_row+1+pi, column=2, value=round(val, 0)).number_format = USD_FMT

auto_width(ws_c4, max_w=16)

# ── Save workbook ──────────────────────────────────────────────────────────────
log(f"EXCEL — saving to {XLSX_PATH} …")
wb.save(XLSX_PATH)
log(f"EXCEL — saved ({XLSX_PATH.stat().st_size/1024:.0f} KB)")

# ══════════════════════════════════════════════════════════════════════════════
# FINAL GIT COMMIT
# ══════════════════════════════════════════════════════════════════════════════
import subprocess

log("GIT — staging outputs …")
try:
    subprocess.run(["git", "add",
                    "outputs/", "src/", "requirements.txt", "README.md", ".gitignore",
                    "data/raw/.gitkeep", "data/processed/.gitkeep"],
                   cwd=str(ROOT), check=True)
    subprocess.run(["git", "commit", "-m",
                    f"Pipeline run {date.today()} — Colombia synthetic MVP pilot"],
                   cwd=str(ROOT), check=True)
    subprocess.run(["git", "push", "-u", "origin", "claude/happy-turing-2osqbg"],
                   cwd=str(ROOT), check=True)
    log("GIT — outputs committed and pushed.")
except Exception as e:
    flag(f"GIT commit/push failed: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# FINAL SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
elapsed_total = time.time() - t0_global
print("\n" + "═"*70)
print("PIPELINE COMPLETE")
print("═"*70)
print(f"  Total elapsed:      {elapsed_total:.1f}s")
print(f"  Grid cells:         {len(grid):,}")
print(f"  Cropland cells:     {n_crop:,}")
print(f"  Plots:              {n_plots:,}")
print(f"  Portfolio slots:    {n_slots:,}")
print(f"  Flags raised:       {len(FLAGS)}")
for f in FLAGS:
    print(f"    • {f}")
print(f"\n  Excel:  {XLSX_PATH}")
for mp in sorted(MAP_DIR.glob("*.png")):
    print(f"  Map:    {mp}")
print("═"*70)
